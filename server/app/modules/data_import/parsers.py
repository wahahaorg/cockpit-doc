from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


SHEETS = {
    "账户余额": {
        "account_code": ("账户编号", "text", True), "account_name": ("账户名称", "text", True),
        "snapshot_date": ("快照日期", "date", True), "available_balance": ("可用余额（元）", "decimal", True),
        "currency": ("币种", "text", False), "restricted_amount": ("受限金额（元）", "decimal", False), "remark": ("备注", "text", False),
    },
    "应收款": {
        "receivable_no": ("应收编号", "text", True), "customer_code": ("客户编号", "text", True),
        "customer_name": ("客户名称", "text", True), "business_ref_no": ("合同/业务编号", "text", False),
        "receivable_amount": ("应收金额（元）", "decimal", True), "agreed_due_date": ("约定到账日期", "date", True),
        "expected_date": ("预计到账日期", "date", False), "owner_code": ("责任人编号", "text", True),
        "owner_name": ("责任人名称", "text", True), "business_line": ("业务线", "text", False),
        "source_status": ("业务状态", "text", True), "remark": ("备注", "text", False),
    },
    "实际回款": {
        "collection_no": ("回款编号", "text", True), "receivable_no": ("应收编号", "text", True),
        "collection_date": ("回款日期", "date", True), "collection_amount": ("回款金额（元）", "decimal", True),
        "currency": ("币种", "text", False), "bank_reference": ("银行流水参考号", "text", False), "remark": ("备注", "text", False),
    },
    "计划支出": {
        "expense_no": ("支出计划编号", "text", True), "expense_name": ("支出名称", "text", True),
        "category": ("支出类别", "text", True), "planned_date": ("计划支出日期", "date", True),
        "planned_amount": ("计划金额（元）", "decimal", True), "owner_code": ("责任人编号", "text", True),
        "owner_name": ("责任人名称", "text", True), "rigidity": ("刚性程度", "text", True),
        "approval_status": ("审批状态", "text", True), "remark": ("备注", "text", False),
    },
}


def _header_name(value: Any) -> str:
    """Normalize the visible V0.1 template header, including required-field marks."""
    return str(value).strip().removesuffix("*").strip() if value is not None else ""


@dataclass
class ParsedRow:
    sheet_name: str
    row_no: int
    raw_data: dict[str, Any]
    normalized_data: dict[str, Any]
    messages: list[dict[str, str]]

    @property
    def status(self) -> str:
        return "error" if any(m["severity"] == "error" for m in self.messages) else ("warning" if self.messages else "valid")


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return value


def _convert(value: Any, kind: str) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if kind == "text":
        return str(value).strip()
    if kind == "date":
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        return date.fromisoformat(str(value).strip())
    if kind == "decimal":
        result = Decimal(str(value).replace(",", "").strip()).quantize(Decimal("0.01"))
        if not result.is_finite():
            raise InvalidOperation
        return result
    return value


def parse_workbook(content: bytes) -> list[ParsedRow]:
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError("无法读取 Excel 文件") from exc
    missing = set(SHEETS) - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"缺少 Sheet：{', '.join(sorted(missing))}")
    result: list[ParsedRow] = []
    for sheet_name, fields in SHEETS.items():
        sheet = workbook[sheet_name]
        required_headers = [spec[0] for spec in fields.values()]
        header_row_no = 0
        headers: list[str] = []
        for row_no, values in enumerate(sheet.iter_rows(values_only=True), start=1):
            candidate = [_header_name(value) for value in values]
            if set(required_headers).issubset(candidate):
                header_row_no = row_no
                headers = candidate
                break
        if not header_row_no:
            raise ValueError(f"{sheet_name} 缺少列：{', '.join(required_headers)}")
        header_map = {name: idx for idx, name in enumerate(headers)}
        rows = sheet.iter_rows(min_row=header_row_no + 1, values_only=True)
        for row_no, values in enumerate(rows, start=header_row_no + 1):
            if not any(v is not None and str(v).strip() for v in values):
                continue
            raw = {headers[i]: _json_value(v) for i, v in enumerate(values) if i < len(headers) and headers[i]}
            normalized: dict[str, Any] = {}
            messages: list[dict[str, str]] = []
            for field, (header, kind, required) in fields.items():
                value = values[header_map[header]] if header_map[header] < len(values) else None
                try:
                    converted = _convert(value, kind)
                except (ValueError, InvalidOperation):
                    converted = None
                    messages.append({"field": field, "code": "invalid_format", "message": f"{header}格式不正确", "severity": "error"})
                if required and converted is None:
                    messages.append({"field": field, "code": "required", "message": f"{header}不能为空", "severity": "error"})
                normalized[field] = converted
            amount_fields = [k for k in normalized if k.endswith("amount") or k == "available_balance"]
            for field in amount_fields:
                if normalized[field] is not None and normalized[field] < 0:
                    messages.append({"field": field, "code": "negative_amount", "message": "金额不得小于 0", "severity": "error"})
            result.append(ParsedRow(sheet_name, row_no, raw, normalized, messages))
    _validate_relations(result)
    return result


def _validate_relations(rows: list[ParsedRow]) -> None:
    receivable_nos = {r.normalized_data.get("receivable_no") for r in rows if r.sheet_name == "应收款" and r.status != "error"}
    seen: dict[tuple[str, str], int] = {}
    keys = {"账户余额": "account_code", "应收款": "receivable_no", "实际回款": "collection_no", "计划支出": "expense_no"}
    for row in rows:
        key = row.normalized_data.get(keys[row.sheet_name])
        identity = (row.sheet_name, str(key))
        if key and identity in seen:
            row.messages.append({"field": keys[row.sheet_name], "code": "duplicate", "message": f"与第 {seen[identity]} 行重复", "severity": "error"})
        elif key:
            seen[identity] = row.row_no
        if row.sheet_name == "实际回款" and row.normalized_data.get("receivable_no") not in receivable_nos:
            row.messages.append({"field": "receivable_no", "code": "reference_not_found", "message": "关联应收编号不存在", "severity": "error"})
        if row.sheet_name == "应收款" and row.normalized_data.get("source_status") not in {"open", "cancelled"}:
            row.messages.append({"field": "source_status", "code": "invalid_enum", "message": "业务状态仅支持 open/cancelled", "severity": "error"})
        if row.sheet_name == "计划支出":
            if row.normalized_data.get("rigidity") not in {"rigid", "deferrable"}:
                row.messages.append({"field": "rigidity", "code": "invalid_enum", "message": "刚性程度仅支持 rigid/deferrable", "severity": "error"})
            if row.normalized_data.get("approval_status") not in {"planned", "pending", "approved", "cancelled"}:
                row.messages.append({"field": "approval_status", "code": "invalid_enum", "message": "审批状态值无效", "severity": "error"})
