import logging
import mimetypes
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import httpx
from openpyxl import load_workbook

from app.core.config import get_settings

logger = logging.getLogger(__name__)


def _extract_settlement_text(path: Path) -> tuple[str, str, str | None]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"# {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [_cell_text(value) for value in row]
                if any(values):
                    lines.append(" | ".join(values))
        return "\n".join(lines), "success", None
    if suffix == ".docx":
        text = _docx_text(path)
        return (text, "success", None) if text.strip() else ("", "failed", "DOCX 未提取到有效文本")
    if suffix == ".pdf":
        text = _pdf_text(path)
        if _pdf_text_has_settlement_detail(text):
            return text, "success", None
        ocr_text, ocr_status, ocr_reason = _ocr_pdf(path)
        if ocr_status == "success":
            return ocr_text, "success", None
        if text.strip():
            return text, "success", None
        return ocr_text, ocr_status, ocr_reason
    if suffix in {".txt", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore"), "success", None
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return _ocr_image(path)
    return "", "failed", "不支持的结算单文件类型"


def _pdf_text(path: Path) -> str:
    try:
        import fitz

        with fitz.open(path) as doc:
            return "\n\n".join(
                f"# 第 {page_index} 页\n{page.get_text().strip()}"
                for page_index, page in enumerate(doc, start=1)
            )
    except Exception:
        return ""


def _pdf_text_has_settlement_detail(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    if len(compact) < 80:
        return False
    has_amount = bool(re.search(r"\d+(?:,\d{3})*(?:\.\d{1,2})?", compact))
    has_settlement_terms = any(term in compact for term in ["合计", "金额", "服务内容", "结算周期", "技术服务费"])
    return has_amount and has_settlement_terms


_ocr_engine = None


def _get_ocr():
    global _ocr_engine
    if _ocr_engine is None:
        from rapidocr_onnxruntime import RapidOCR

        _ocr_engine = RapidOCR()
    return _ocr_engine


def _ocr_with_service(path: Path) -> tuple[str, str, str | None]:
    settings = get_settings()
    if not settings.ocr_service_url:
        return "", "failed", "未配置 OCR 服务地址"
    try:
        with path.open("rb") as file:
            response = httpx.post(
                settings.ocr_service_url,
                files={"file": (path.name, file, mimetypes.guess_type(path.name)[0] or "application/octet-stream")},
                timeout=settings.ocr_service_timeout_seconds,
            )
        response.raise_for_status()
        payload = response.json()
        text = payload.get("text") if isinstance(payload, dict) else None
        if not text and isinstance(payload, dict) and isinstance(payload.get("lines"), list):
            text = "\n".join(
                str(item.get("text") or "").strip()
                for item in payload["lines"]
                if isinstance(item, dict) and item.get("text")
            )
        if isinstance(text, str) and text.strip():
            return text.strip(), "success", None
        return "", "needs_ocr", "PaddleOCR 服务未识别到有效文本"
    except Exception as exc:
        logger.warning("income_reconciliation_remote_ocr_failed file=%s url=%s error=%s", path.name, settings.ocr_service_url, exc)
        return "", "failed", f"PaddleOCR 服务调用失败：{exc}"


def _ocr_with_local_engine(path: Path) -> tuple[str, str, str | None]:
    try:
        result, _elapse = _get_ocr()(str(path))
        if result is None:
            return "", "needs_ocr", "OCR 未识别到有效文本"
        return "\n".join(item[1] for item in result), "success", None
    except ImportError:
        return "", "needs_ocr", "OCR 引擎未安装，请安装 rapidocr-onnxruntime"
    except Exception as exc:
        return "", "failed", f"OCR 识别失败：{exc}"


def _ocr_image(path: Path) -> tuple[str, str, str | None]:
    """对单张图片执行 OCR，返回 (text, status, error_message)。"""
    settings = get_settings()
    if settings.ocr_service_url:
        remote_result = _ocr_with_service(path)
        if remote_result[1] == "success" or not settings.ocr_local_fallback_enabled:
            return remote_result
    return _ocr_with_local_engine(path)


def _ocr_pdf(path: Path) -> tuple[str, str, str | None]:
    """将 PDF 每页渲染为图片后执行 OCR，返回 (text, status, error_message)。"""
    try:
        import fitz
    except ImportError:
        return "", "needs_ocr", "缺少 pymupdf，无法渲染扫描 PDF"

    try:
        all_lines: list[str] = []
        failed_reasons: list[str] = []
        with fitz.open(path) as doc:
            pages = list(doc)
            if not pages:
                return "", "failed", "PDF 文件没有可识别页面"

            for page_idx, page in enumerate(pages, start=1):
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                tmp = path.with_name(f".ocr_tmp_{path.stem}_p{page_idx}.png")
                pixmap.save(tmp)
                try:
                    page_text, page_status, page_reason = _ocr_image(tmp)
                    if page_status == "success" and page_text:
                        all_lines.append(f"# 第 {page_idx} 页\n{page_text}")
                        all_lines.append("")
                    elif page_reason:
                        failed_reasons.append(f"第 {page_idx} 页：{page_reason}")
                finally:
                    tmp.unlink(missing_ok=True)

        text = "\n".join(all_lines).strip()
        if text:
            return text, "success", None
        if failed_reasons:
            return "", "failed", "；".join(failed_reasons)
        return "", "needs_ocr", "PDF 经 OCR 后未提取到有效文本"
    except Exception as exc:
        return "", "failed", f"PDF OCR 失败：{exc}"


def _docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        root = ElementTree.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        return "\n".join(node.text for node in root.findall(".//w:t", ns) if node.text)
    except Exception:
        return ""


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
