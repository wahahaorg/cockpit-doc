import json
import logging
import re
import shutil
import time
import uuid
import zipfile
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from fastapi import UploadFile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.core.config import get_settings
from app.core.exceptions import AppError
from app.modules.income_reconciliation.company_name import CompanyMatchRelation, compare_company_names, normalize_company_name

logger = logging.getLogger(__name__)

JOB_ROOT = Path(get_settings().income_reconciliation_storage_dir) / "income-reconciliation" / "jobs"
MONEY_TOLERANCE = Decimal("0.01")


def create_job(
    invoice_file: UploadFile,
    cashflow_file: UploadFile,
    settlement_files: list[UploadFile],
    period_start: str | None,
    period_end: str | None,
) -> dict[str, Any]:
    if not invoice_file.filename or not invoice_file.filename.lower().endswith(".xlsx"):
        raise AppError("invalid_invoice_file", "开票明细仅支持 .xlsx 文件", 415)
    if not cashflow_file.filename or not cashflow_file.filename.lower().endswith(".xlsx"):
        raise AppError("invalid_cashflow_file", "服务收支明细仅支持 .xlsx 文件", 415)
    if not settlement_files:
        raise AppError("missing_settlement_files", "请至少上传一个结算单文件", 422)

    job_id = f"job_{datetime.now():%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:8]}"
    job_dir = _job_dir(job_id)
    (job_dir / "uploads" / "settlements").mkdir(parents=True, exist_ok=True)
    (job_dir / "parsed").mkdir(parents=True, exist_ok=True)
    (job_dir / "extracted" / "texts").mkdir(parents=True, exist_ok=True)
    (job_dir / "extracted" / "ai_extracts").mkdir(parents=True, exist_ok=True)
    (job_dir / "result").mkdir(parents=True, exist_ok=True)

    _save_upload(invoice_file, job_dir / "uploads" / _safe_name(invoice_file.filename))
    _save_upload(cashflow_file, job_dir / "uploads" / _safe_name(cashflow_file.filename))
    saved_settlements = []
    for file in settlement_files:
        if not file.filename:
            continue
        target = job_dir / "uploads" / "settlements" / _safe_name(file.filename)
        _save_upload(file, target)
        saved_settlements.append(target.name)

    metadata = {
        "jobId": job_id,
        "status": "uploaded",
        "periodStart": period_start,
        "periodEnd": period_end,
        "createdAt": datetime.now().isoformat(),
        "invoiceFile": _safe_name(invoice_file.filename),
        "cashflowFile": _safe_name(cashflow_file.filename),
        "settlementFiles": saved_settlements,
        "summary": None,
        "downloadUrl": None,
    }
    _write_json(job_dir / "job.json", metadata)
    _write_json(job_dir / "parsed" / "files.json", [])
    _reset_progress_events(job_dir)
    _append_progress_event(job_dir, "job_created", "收入核对任务已创建", progress=0)
    return get_job(job_id)


def parse_job(job_id: str) -> dict[str, Any]:
    started_at = time.monotonic()
    job_dir = _require_job(job_id)
    metadata = _read_json(job_dir / "job.json")
    should_reset_events = metadata.get("status") != "parsing"
    if should_reset_events:
        _reset_progress_events(job_dir)
    logger.info(
        "income_reconciliation_parse_started job_id=%s invoice_file=%s cashflow_file=%s settlement_count=%d",
        job_id,
        metadata.get("invoiceFile"),
        metadata.get("cashflowFile"),
        len(metadata.get("settlementFiles", [])),
    )
    metadata["status"] = "parsing"
    _write_json(job_dir / "job.json", metadata)
    _append_progress_event(job_dir, "job_started", "开始解析收入核对任务", progress=5)

    files = _read_json(job_dir / "parsed" / "files.json", []) if not should_reset_events else _initial_file_statuses(metadata)
    if not files:
        files = _initial_file_statuses(metadata)
    _write_json(job_dir / "parsed" / "files.json", files)
    invoice_rows: list[dict[str, Any]] = []
    cashflow_rows: list[dict[str, Any]] = []
    settlement_rows: list[dict[str, Any]] = []

    invoice_path = job_dir / "uploads" / metadata["invoiceFile"]
    cashflow_path = job_dir / "uploads" / metadata["cashflowFile"]
    try:
        _replace_file_status(files, _file_status("file_invoice", invoice_path.name, "invoice_excel", "parsing", 0, 0, 0, None))
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_started", "正在读取开票明细 Excel", file_id="file_invoice", file_name=invoice_path.name, stage="excel_parse", progress=12)
        invoice_started_at = time.monotonic()
        invoice_rows = _parse_invoice_excel(invoice_path)
        logger.info(
            "income_reconciliation_invoice_parsed job_id=%s file=%s rows=%d elapsed_ms=%d",
            job_id,
            invoice_path.name,
            len(invoice_rows),
            int((time.monotonic() - invoice_started_at) * 1000),
        )
        invoice_status = _file_status("file_invoice", invoice_path.name, "invoice_excel", "success", len(invoice_rows), sum(1 for row in invoice_rows if row["isEffective"]), 1, None)
        _replace_file_status(files, invoice_status)
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_done", "开票明细 Excel 读取完成", file_id="file_invoice", file_name=invoice_path.name, stage="excel_parse", parsed_rows=len(invoice_rows), valid_rows=invoice_status["validRows"], progress=25)
    except Exception as exc:
        logger.exception("income_reconciliation_invoice_failed job_id=%s file=%s", job_id, invoice_path.name)
        _replace_file_status(files, _file_status("file_invoice", invoice_path.name, "invoice_excel", "failed", 0, 0, 0, str(exc)))
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_failed", "开票明细 Excel 读取失败", file_id="file_invoice", file_name=invoice_path.name, stage="excel_parse", reason=str(exc), progress=25)

    try:
        _replace_file_status(files, _file_status("file_cashflow", cashflow_path.name, "cashflow_excel", "parsing", 0, 0, 0, None))
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_started", "正在读取服务收支明细 Excel", file_id="file_cashflow", file_name=cashflow_path.name, stage="excel_parse", progress=30)
        cashflow_started_at = time.monotonic()
        cashflow_rows = _parse_cashflow_excel(cashflow_path)
        logger.info(
            "income_reconciliation_cashflow_parsed job_id=%s file=%s rows=%d elapsed_ms=%d",
            job_id,
            cashflow_path.name,
            len(cashflow_rows),
            int((time.monotonic() - cashflow_started_at) * 1000),
        )
        cashflow_status = _file_status("file_cashflow", cashflow_path.name, "cashflow_excel", "success", len(cashflow_rows), len(cashflow_rows), 1, None)
        _replace_file_status(files, cashflow_status)
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_done", "服务收支明细 Excel 读取完成", file_id="file_cashflow", file_name=cashflow_path.name, stage="excel_parse", parsed_rows=len(cashflow_rows), valid_rows=cashflow_status["validRows"], progress=40)
    except Exception as exc:
        logger.exception("income_reconciliation_cashflow_failed job_id=%s file=%s", job_id, cashflow_path.name)
        _replace_file_status(files, _file_status("file_cashflow", cashflow_path.name, "cashflow_excel", "failed", 0, 0, 0, str(exc)))
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_failed", "服务收支明细 Excel 读取失败", file_id="file_cashflow", file_name=cashflow_path.name, stage="excel_parse", reason=str(exc), progress=40)

    settlement_count = len(metadata.get("settlementFiles", []))
    for index, name in enumerate(metadata.get("settlementFiles", []), start=1):
        file_id = f"file_settlement_{index:03d}"
        path = job_dir / "uploads" / "settlements" / name
        progress_base = 45 + int((index - 1) / max(settlement_count, 1) * 40)
        _replace_file_status(files, _file_status(file_id, path.name, _settlement_type(path), "parsing", 0, 0, 0, None))
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_started", "正在解析结算单文件", file_id=file_id, file_name=path.name, stage="settlement_parse", progress=progress_base)
        parsed, status = _parse_settlement(path, file_id, job_dir, progress_base=progress_base)
        settlement_rows.extend(parsed)
        _replace_file_status(files, status)
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "file_done" if status["parseStatus"] in {"success", "warning"} else "file_failed", "结算单文件解析完成" if status["parseStatus"] in {"success", "warning"} else "结算单文件解析失败", file_id=file_id, file_name=path.name, stage="settlement_parse", parsed_rows=status["parsedRows"], valid_rows=status["validRows"], confidence=status["confidence"], reason=status.get("errorReason"), progress=min(progress_base + 12, 88))

    _write_json(job_dir / "parsed" / "files.json", files)
    _write_json(job_dir / "parsed" / "invoices.json", invoice_rows)
    _write_json(job_dir / "parsed" / "cashflows.json", cashflow_rows)
    _write_json(job_dir / "parsed" / "settlements.json", settlement_rows)

    metadata["status"] = "parse_failed" if any(file["parseStatus"] == "failed" for file in files[:2]) else "parsed"
    _write_json(job_dir / "job.json", metadata)
    _append_progress_event(job_dir, "job_done", "收入核对解析完成" if metadata["status"] == "parsed" else "收入核对解析失败，请查看文件原因", status=metadata["status"], progress=100)
    logger.info(
        "income_reconciliation_parse_finished job_id=%s status=%s invoice_rows=%d cashflow_rows=%d settlement_rows=%d elapsed_ms=%d",
        job_id,
        metadata["status"],
        len(invoice_rows),
        len(cashflow_rows),
        len(settlement_rows),
        int((time.monotonic() - started_at) * 1000),
    )
    return get_job(job_id)


def prepare_parse_job(job_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    metadata = _read_json(job_dir / "job.json")
    metadata["status"] = "parsing"
    _write_json(job_dir / "job.json", metadata)
    _reset_progress_events(job_dir)
    _write_json(job_dir / "parsed" / "files.json", _initial_file_statuses(metadata))
    _append_progress_event(job_dir, "job_queued", "解析任务已进入后台队列", progress=1)
    return get_job(job_id)


def generate_job(job_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    metadata = _read_json(job_dir / "job.json")
    metadata["status"] = "generating"
    _write_json(job_dir / "job.json", metadata)

    invoices = _read_json(job_dir / "parsed" / "invoices.json", [])
    cashflows = _read_json(job_dir / "parsed" / "cashflows.json", [])
    settlements = _read_json(job_dir / "parsed" / "settlements.json", [])
    files = _read_json(job_dir / "parsed" / "files.json", [])
    reconciliation = _reconcile(invoices, cashflows, settlements)
    _write_json(job_dir / "result" / "reconciliation.json", reconciliation)
    _export_excel(job_dir / "result" / "reconciliation.xlsx", reconciliation, files)

    metadata["status"] = "generated"
    metadata["summary"] = reconciliation["summary"]
    metadata["downloadUrl"] = f"/api/v1/income-reconciliation/jobs/{job_id}/download"
    _write_json(job_dir / "job.json", metadata)
    return get_job(job_id)


def get_job(job_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    metadata = _read_json(job_dir / "job.json")
    metadata["files"] = _read_json(job_dir / "parsed" / "files.json", [])
    return metadata


def get_progress_events(job_id: str) -> list[dict[str, Any]]:
    job_dir = _require_job(job_id)
    path = job_dir / "progress.jsonl"
    if not path.exists():
        return []
    events = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            logger.warning("income_reconciliation_progress_event_invalid job_id=%s line=%s", job_id, line[:200])
    return events


def get_file_result(job_id: str, file_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    files = _read_json(job_dir / "parsed" / "files.json", [])
    item = next((file for file in files if file["fileId"] == file_id), None)
    if not item:
        raise AppError("file_not_found", "解析文件不存在", 404)
    return {
        "fileId": file_id,
        "fileName": item["fileName"],
        "rawText": _read_text(job_dir / item.get("textPath", "")),
        "aiExtractedJson": _read_json(job_dir / item.get("aiExtractPath", ""), None) if item.get("aiExtractPath") else None,
        "standardJson": _standard_json_for_file(job_dir, file_id, item["fileType"]),
    }


def retry_settlement_ai(job_id: str, file_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    files = _read_json(job_dir / "parsed" / "files.json", [])
    item = next((file for file in files if file.get("fileId") == file_id), None)
    if not item:
        raise AppError("file_not_found", "解析文件不存在", 404)
    if not str(item.get("fileType", "")).startswith("settlement_"):
        raise AppError("not_settlement_file", "只有结算单文件支持 AI 重试", 409)
    text = _read_text(job_dir / item.get("textPath", ""))
    if not text.strip():
        raise AppError("ocr_text_not_found", "该结算单没有可用于重试的 OCR 文本", 409)

    retry_count = int(item.get("aiRetryCount") or 0) + 1
    _append_progress_event(job_dir, "ai_retry_started", "正在重新调用 AI 抽取结算单", file_id=file_id, file_name=item["fileName"], stage="ai_extract")
    ai_result = _extract_settlement_with_ai(text, item["fileName"])
    ai_rel = item.get("aiExtractPath") or f"extracted/ai_extracts/{file_id}.json"
    _write_json(job_dir / ai_rel, ai_result)
    item.update({"aiExtractPath": ai_rel, "aiRetryCount": retry_count, "aiLastTriedAt": ai_result.get("triedAt")})

    if ai_result.get("status") == "failed":
        reason = ai_result.get("error") or "AI 抽取失败"
        item.update({"parseStatus": "failed", "parsedRows": 0, "validRows": 0, "confidence": 0, "errorReason": f"AI 抽取失败：{reason}", "aiStatus": "failed", "aiError": reason})
        _write_json(job_dir / "parsed" / "files.json", files)
        _append_progress_event(job_dir, "ai_retry_failed", "AI 重试失败", file_id=file_id, file_name=item["fileName"], stage="ai_extract", reason=reason)
        return get_job(job_id)

    records = _standardize_settlement_records(ai_result, file_id, item["fileName"], text)
    parse_status = "success" if records and all(record["parseStatus"] == "success" for record in records) else "warning"
    item.update({
        "parseStatus": parse_status,
        "parsedRows": len(records),
        "validRows": sum(record["parseStatus"] == "success" for record in records),
        "confidence": max((record["confidence"] for record in records), default=0),
        "errorReason": None if parse_status == "success" else "存在低置信度或缺失字段",
        "aiStatus": "success",
        "aiError": None,
    })
    settlements = _read_json(job_dir / "parsed" / "settlements.json", [])
    settlements = [record for record in settlements if record.get("fileId") != file_id]
    settlements.extend(records)
    _write_json(job_dir / "parsed" / "settlements.json", settlements)
    _write_json(job_dir / "parsed" / "files.json", files)
    _invalidate_generated_result(job_dir)
    _append_progress_event(job_dir, "ai_retry_done", "AI 重试成功，结算单字段已更新", file_id=file_id, file_name=item["fileName"], stage="ai_extract", parsed_rows=len(records), valid_rows=item["validRows"], confidence=item["confidence"])
    return get_job(job_id)


def retry_failed_settlement_ai(job_id: str) -> dict[str, Any]:
    job_dir = _require_job(job_id)
    files = _read_json(job_dir / "parsed" / "files.json", [])
    failed_ids = [item["fileId"] for item in files if item.get("aiStatus") == "failed"]
    if not failed_ids:
        raise AppError("no_failed_ai_files", "当前没有需要重试的 AI 失败文件", 409)
    for file_id in failed_ids:
        retry_settlement_ai(job_id, file_id)
    return get_job(job_id)


def _invalidate_generated_result(job_dir: Path) -> None:
    metadata = _read_json(job_dir / "job.json", {})
    if metadata.get("status") == "generated" or metadata.get("downloadUrl"):
        for name in ("reconciliation.json", "reconciliation.xlsx"):
            (job_dir / "result" / name).unlink(missing_ok=True)
        metadata.update({"status": "parsed", "summary": None, "downloadUrl": None})
        _write_json(job_dir / "job.json", metadata)


def download_path(job_id: str) -> Path:
    path = _require_job(job_id) / "result" / "reconciliation.xlsx"
    if not path.exists():
        raise AppError("excel_not_generated", "核对表尚未生成", 404)
    return path


def _parse_invoice_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws, header_row, headers = _find_sheet_header(wb, ["数电发票号码", "购买方名称", "开票日期"], "发票基础信息")
    rows = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        raw = _row_dict(ws, row_no, headers)
        if not any(raw.values()):
            continue
        invoice_no = _text(raw.get("数电发票号码"))
        customer_name = _text(raw.get("购买方名称"))
        invoice_date = _date_text(raw.get("开票日期"))
        amount = _money(raw.get("价税合计") if raw.get("价税合计") is not None else raw.get("金额"))
        status = _text(raw.get("发票状态")) or "未知"
        positive = _text(raw.get("是否正数发票"))
        is_effective = status == "正常" and positive == "是"
        invalid_reason = None
        if not is_effective:
            if "红冲" in status or positive == "否":
                invalid_reason = "红字/红冲发票，不计入确认收入"
            elif "作废" in status:
                invalid_reason = "作废发票，不计入确认收入"
            else:
                invalid_reason = "非正常或非正数发票，不计入确认收入"
        rows.append({
            "sourceFile": path.name,
            "sourceSheet": ws.title,
            "sourceRowNo": row_no,
            "invoiceNo": invoice_no,
            "customerName": customer_name,
            "invoiceDate": invoice_date,
            "invoiceMonth": invoice_date[:7] if invoice_date else None,
            "invoiceAmount": _float(amount),
            "isEffective": is_effective,
            "invalidReason": invalid_reason,
            "rawRow": raw,
        })
    return rows


def _parse_cashflow_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    ws, header_row, headers = _find_sheet_header(wb, ["日期", "往来单位", "收入(借方)"], None)
    rows = []
    for row_no in range(header_row + 1, ws.max_row + 1):
        raw = _row_dict(ws, row_no, headers)
        if not any(raw.values()):
            continue
        amount = _money(raw.get("收入(借方)"))
        if amount <= 0:
            continue
        rows.append({
            "sourceFile": path.name,
            "sourceSheet": ws.title,
            "sourceRowNo": row_no,
            "customerName": _text(raw.get("往来单位")),
            "transactionDate": _date_text(raw.get("日期")),
            "receivedAmount": _float(amount),
            "summary": _text(raw.get("摘要")),
            "rawRow": raw,
        })
    return rows


def _parse_settlement(path: Path, file_id: str, job_dir: Path, progress_base: int = 50) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    started_at = time.monotonic()
    file_type = _settlement_type(path)
    logger.info(
        "income_reconciliation_settlement_parse_started file_id=%s file=%s file_type=%s",
        file_id,
        path.name,
        file_type,
    )
    _append_progress_event(job_dir, "stage_started", "正在提取结算单文本", file_id=file_id, file_name=path.name, stage="text_extract", progress=progress_base + 2)
    text, text_status, text_reason = _extract_settlement_text(path)
    text_rel = f"extracted/texts/{file_id}.txt"
    _write_text(job_dir / text_rel, text)
    logger.info(
        "income_reconciliation_settlement_text_extracted file_id=%s file=%s status=%s text_chars=%d elapsed_ms=%d",
        file_id,
        path.name,
        text_status,
        len(text),
        int((time.monotonic() - started_at) * 1000),
    )
    if text_status != "success":
        logger.warning(
            "income_reconciliation_settlement_text_failed file_id=%s file=%s reason=%s",
            file_id,
            path.name,
            text_reason,
        )
        status = _file_status(file_id, path.name, file_type, text_status, 0, 0, 0, text_reason)
        status["textPath"] = text_rel
        _append_progress_event(job_dir, "stage_failed", "结算单文本提取失败", file_id=file_id, file_name=path.name, stage="text_extract", reason=text_reason, progress=progress_base + 6)
        return [], status

    _append_progress_event(job_dir, "stage_done", "结算单文本提取完成", file_id=file_id, file_name=path.name, stage="text_extract", text_chars=len(text), progress=progress_base + 6)
    _append_progress_event(job_dir, "ai_started", "AI 正在抽取客户、周期、结算金额", file_id=file_id, file_name=path.name, stage="ai_extract", progress=progress_base + 8)
    ai_result = _extract_settlement_with_ai(text, path.name)
    ai_rel = f"extracted/ai_extracts/{file_id}.json"
    _write_json(job_dir / ai_rel, ai_result)
    if ai_result.get("status") == "failed":
        reason = ai_result.get("error") or "AI 抽取失败"
        status = _file_status(file_id, path.name, file_type, "failed", 0, 0, 0, f"AI 抽取失败：{reason}")
        status.update({
            "textPath": text_rel,
            "aiExtractPath": ai_rel,
            "aiStatus": "failed",
            "aiError": reason,
            "aiRetryCount": 0,
            "aiLastTriedAt": ai_result.get("triedAt"),
        })
        _append_progress_event(job_dir, "ai_failed", "AI 抽取失败，可在文件列表中重试", file_id=file_id, file_name=path.name, stage="ai_extract", reason=reason, progress=progress_base + 10)
        return [], status
    _append_progress_event(job_dir, "ai_done", "AI 抽取完成，正在标准化字段", file_id=file_id, file_name=path.name, stage="ai_extract", model=ai_result.get("model"), progress=progress_base + 10)
    records = _standardize_settlement_records(ai_result, file_id, path.name, text)
    parse_status = "success" if records and all(item["parseStatus"] == "success" for item in records) else "warning"
    status = _file_status(file_id, path.name, file_type, parse_status, len(records), sum(r["parseStatus"] == "success" for r in records), max((r["confidence"] for r in records), default=0), None if parse_status == "success" else "存在低置信度或缺失字段")
    status.update({"textPath": text_rel, "aiExtractPath": ai_rel, "aiStatus": "success", "aiError": None, "aiRetryCount": 0, "aiLastTriedAt": ai_result.get("triedAt")})
    _append_progress_event(job_dir, "stage_done", "结算单字段标准化完成", file_id=file_id, file_name=path.name, stage="standardize", parsed_rows=len(records), valid_rows=status["validRows"], confidence=status["confidence"], reason=status.get("errorReason"), progress=progress_base + 11)
    logger.info(
        "income_reconciliation_settlement_parse_finished file_id=%s file=%s status=%s model=%s records=%d elapsed_ms=%d",
        file_id,
        path.name,
        parse_status,
        ai_result.get("model"),
        len(records),
        int((time.monotonic() - started_at) * 1000),
    )
    return records, status


def _standardize_settlement_records(ai_result: dict[str, Any], file_id: str, file_name: str, text: str) -> list[dict[str, Any]]:
    records = []
    for record in ai_result.get("records", []):
        confidence = float(record.get("confidence") or 0)
        missing = record.get("missingFields") or []
        parse_status = "success" if confidence >= 0.8 and not missing else "warning"
        parse_reason = None if parse_status == "success" else "AI 置信度低或关键字段缺失，需人工确认"
        records.append({
            "fileId": file_id,
            "sourceFile": file_name,
            "sourceType": Path(file_name).suffix.lower().lstrip(".") or "file",
            "customerName": _text(record.get("customerName")) or None,
            "settlementPeriod": _month_text(record.get("settlementPeriod")),
            "settlementAmount": _float(_money(record.get("settlementAmount"))) if record.get("settlementAmount") not in (None, "") else None,
            "confidence": confidence,
            "parseStatus": parse_status,
            "parseReason": parse_reason,
            "rawText": text[:4000],
        })
    return records


def _extract_settlement_with_ai(text: str, file_name: str) -> dict[str, Any]:
    settings = get_settings()
    prompt = _settlement_ai_prompt(text)
    tried_at = datetime.now().isoformat()
    try:
        from app.modules.ai.client import ai_available, get_chat_model

        ai_enabled = ai_available()
        if settings.income_reconciliation_ai_enabled and ai_enabled:
            started_at = time.monotonic()
            logger.info(
                "income_reconciliation_ai_call_started file=%s model=%s base_url=%s prompt_chars=%d text_chars=%d timeout_seconds=%d",
                file_name,
                settings.ai_model,
                settings.ollama_base_url,
                len(prompt),
                len(text),
                settings.ai_timeout_seconds,
            )
            response = get_chat_model().invoke(prompt)
            content = response.content if hasattr(response, "content") else str(response)
            parsed = _json_from_text(content)
            if isinstance(parsed, dict) and isinstance(parsed.get("records"), list):
                parsed["model"] = settings.ai_model
                parsed["status"] = "success"
                parsed["triedAt"] = tried_at
                logger.info(
                    "income_reconciliation_ai_call_finished file=%s model=%s records=%d response_chars=%d elapsed_ms=%d",
                    file_name,
                    settings.ai_model,
                    len(parsed["records"]),
                    len(content),
                    int((time.monotonic() - started_at) * 1000),
                )
                return parsed
            logger.warning(
                "income_reconciliation_ai_invalid_json file=%s model=%s response_chars=%d",
                file_name,
                settings.ai_model,
                len(content),
            )
            return {"status": "failed", "records": [], "model": settings.ai_model, "error": "AI 返回内容不是有效的结算单 JSON", "triedAt": tried_at}
        else:
            logger.info(
                "income_reconciliation_ai_skipped file=%s income_ai_enabled=%s ai_enabled=%s",
                file_name,
                settings.income_reconciliation_ai_enabled,
                ai_enabled,
            )
            return {"status": "failed", "records": [], "model": settings.ai_model, "error": "AI 服务未启用或当前不可用", "triedAt": tried_at}
    except Exception as exc:
        logger.exception(
            "income_reconciliation_ai_call_failed file=%s model=%s base_url=%s",
            file_name,
            settings.ai_model,
            settings.ollama_base_url,
        )
        return {"status": "failed", "records": [], "model": settings.ai_model, "error": str(exc), "triedAt": tried_at}
    return {"status": "failed", "records": [], "model": settings.ai_model, "error": "AI 抽取失败", "triedAt": tried_at}


def _settlement_ai_prompt(text: str) -> str:
    return (
        "你是财务结算单解析助手。请从结算单 OCR/文本中抽取结构化字段，只输出 JSON，不要解释。\n"
        "目标字段：customerName=收款方/服务提供方公司名称，settlementPeriod=结算周期 YYYY-MM，settlementAmount=本结算单最终应结算金额，confidence=0到1，missingFields=缺失字段数组。\n"
        "抽取规则：\n"
        "1. customerName 只提取本次结算的收款方、乙方或服务提供方公司名称。\n"
        "2. 不要把付款方、甲方、采购方或付款方开票信息中的公司名称作为 customerName；如果正文同时出现付款方和收款方，以收款方为准。\n"
        "3. settlementAmount 表示本结算单最终应结算的总金额。OCR 文本可能存在换行错乱、列顺序丢失、标题与数值分离，请结合上下文恢复字段与数值的对应关系。\n"
        "4. 识别金额候选时，优先级依次为：明确标注的合计/总计/应结算金额，其次是可通过数量乘以单价验证的总价，再次是服务费用或含税总额，最后才是其他金额。\n"
        "5. 当存在调用次数、数量、单价、总价等字段时，请使用数量乘以单价对总价进行交叉验证；计算结果与某个金额候选一致时，应提高该候选的可信度。\n"
        "6. 不要仅因为金额出现在项目行附近就排除它。如果该金额同时满足合计标签、总价字段或数量乘以单价的计算关系，应将其作为 settlementAmount。不要把调用次数、数量或单价本身当作结算金额。\n"
        "7. 如果存在多个金额候选，请在内部完成候选比较和算术核验后选择证据最充分的一项，不要简单选择 OCR 文本中最后出现的数字。\n"
        "8. settlementPeriod 只从正文明确的结算周期提取。\n"
        "9. 不要编造缺失信息；找不到的字段填 null。金额输出数字，不要带人民币、元、逗号或其他单位。\n"
        "10. 如果一个文件里有多张独立结算单，records 输出多条；如果只是同一张结算单的多行服务项目，只输出一条。\n"
        "只输出最终 JSON，不要输出分析过程。输出格式：{\"records\":[{\"customerName\":null,\"settlementPeriod\":null,\"settlementAmount\":null,\"confidence\":0.0,\"missingFields\":[]}]}\n"
        f"文本：\n{text[:12000]}"
    )


def _reconcile(invoices: list[dict[str, Any]], cashflows: list[dict[str, Any]], settlements: list[dict[str, Any]]) -> dict[str, Any]:
    items = []
    used_cashflow: set[int] = set()
    used_settlement: set[int] = set()

    for invoice in invoices:
        if not invoice.get("isEffective"):
            items.append(_result_item(invoice, None, None, "发票已红冲", invoice.get("invalidReason"), True))
            continue
        cashflow_index, cashflow, cashflow_relation, cashflow_issue = _find_match(invoice, cashflows, "receivedAmount", used_cashflow, candidate_label="到账")
        if cashflow_index is not None:
            used_cashflow.add(cashflow_index)
        settlement_index, settlement, settlement_relation, settlement_issue = _find_match(
            invoice,
            settlements,
            "settlementAmount",
            used_settlement,
            alternate_names=[cashflow.get("customerName")] if cashflow else None,
            candidate_label="结算单",
        )
        if settlement_index is not None:
            used_settlement.add(settlement_index)
        status = "已确认已到账"
        reason = None
        manual = False
        amounts = [Decimal(str(invoice.get("invoiceAmount") or 0))]
        if settlement:
            amounts.append(Decimal(str(settlement.get("settlementAmount") or 0)))
        if cashflow:
            amounts.append(Decimal(str(cashflow.get("receivedAmount") or 0)))
        if any(abs(amount - amounts[0]) > MONEY_TOLERANCE for amount in amounts[1:]):
            status, reason, manual = "金额异常", "结算单、发票、到账金额不一致", True
        elif not cashflow:
            status, reason, manual = "已确认未到账", cashflow_issue or "有有效发票但未匹配到到账记录", True
        elif not settlement:
            status, reason, manual = "资料缺失待确认", settlement_issue or "有有效发票但未匹配到结算单", True
        items.append(_result_item(invoice, settlement, cashflow, status, reason, manual, cashflow_relation, settlement_relation))

    for index, cashflow in enumerate(cashflows):
        if index in used_cashflow:
            continue
        items.append(_result_item(None, None, cashflow, "未确认已到账", "有到账但未匹配到有效发票", True))

    for index, settlement in enumerate(settlements):
        if index in used_settlement:
            continue
        items.append(_result_item(None, settlement, None, "资料缺失待确认", "有结算单但未匹配到有效发票", True))

    confirmed = sum(Decimal(str(i.get("invoiceAmount") or 0)) for i in items if i.get("invoiceAmount") and i["status"] != "发票已红冲")
    received = sum(Decimal(str(i.get("receivedAmount") or 0)) for i in items if i.get("invoiceAmount") and i.get("receivedAmount") and i["status"] != "发票已红冲")
    summary = {
        "confirmedRevenue": _float(confirmed),
        "receivedAmount": _float(received),
        "unreceivedAmount": _float(confirmed - received),
        "normalCount": sum(i["status"] == "已确认已到账" for i in items),
        "abnormalCount": sum(i["status"] not in {"已确认已到账", "发票已红冲"} for i in items),
        "invalidInvoiceCount": sum(i["status"] == "发票已红冲" for i in items),
        "manualCheckRequiredCount": sum(bool(i["manualCheckRequired"]) for i in items),
    }
    return {"summary": summary, "items": items}


def _result_item(
    invoice: dict[str, Any] | None,
    settlement: dict[str, Any] | None,
    cashflow: dict[str, Any] | None,
    status: str,
    reason: str | None,
    manual: bool,
    cashflow_relation: str | None = None,
    settlement_relation: str | None = None,
) -> dict[str, Any]:
    return {
        "customerName": (invoice or settlement or cashflow or {}).get("customerName"),
        "invoiceNo": invoice.get("invoiceNo") if invoice else None,
        "invoiceDate": invoice.get("invoiceDate") if invoice else None,
        "revenueMonth": invoice.get("invoiceMonth") if invoice else None,
        "invoiceAmount": invoice.get("invoiceAmount") if invoice else None,
        "invoiceSourceFile": invoice.get("sourceFile") if invoice else None,
        "invoiceSourceRowNo": invoice.get("sourceRowNo") if invoice else None,
        "settlementFile": settlement.get("sourceFile") if settlement else None,
        "settlementFileId": settlement.get("fileId") if settlement else None,
        "settlementPeriod": settlement.get("settlementPeriod") if settlement else None,
        "settlementAmount": settlement.get("settlementAmount") if settlement else None,
        "settlementConfidence": settlement.get("confidence") if settlement else None,
        "cashflowDate": cashflow.get("transactionDate") if cashflow else None,
        "cashflowSourceFile": cashflow.get("sourceFile") if cashflow else None,
        "cashflowSourceRowNo": cashflow.get("sourceRowNo") if cashflow else None,
        "receivedAmount": cashflow.get("receivedAmount") if cashflow else None,
        "status": status,
        "abnormalReason": reason,
        "manualCheckRequired": manual,
        "abnormalStage": _abnormal_stage(status),
        "cashflowCompanyMatch": cashflow_relation,
        "settlementCompanyMatch": settlement_relation,
        "matchRule": "amount_then_company_v1" if invoice else None,
    }


def _export_excel(path: Path, reconciliation: dict[str, Any], files: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "老板卡片"
    summary_rows = [["指标", "金额/数量"], ["确认收入总额", reconciliation["summary"]["confirmedRevenue"]], ["已到账金额", reconciliation["summary"]["receivedAmount"]], ["未到账金额", reconciliation["summary"]["unreceivedAmount"]], ["正常笔数", reconciliation["summary"]["normalCount"]], ["异常笔数", reconciliation["summary"]["abnormalCount"]], ["红冲/无效发票笔数", reconciliation["summary"]["invalidInvoiceCount"]], ["待财务确认笔数", reconciliation["summary"]["manualCheckRequiredCount"]]]
    for row in summary_rows:
        ws.append(row)
    _style_header(ws, 1)

    headers = ["客户/项目", "发票号", "开票日期", "收入归属月份", "发票金额", "发票来源文件", "发票来源行号", "结算单文件", "结算周期", "结算金额", "结算单置信度", "到账日期", "到账金额", "到账来源文件", "到账来源行号", "系统判断状态", "异常原因", "异常来源环节", "倒查文件ID", "待人工确认事项"]
    ws = wb.create_sheet("收入链路核对表")
    ws.append(headers)
    _style_header(ws, 1)
    for item in reconciliation["items"]:
        ws.append([item.get("customerName"), item.get("invoiceNo"), item.get("invoiceDate"), item.get("revenueMonth"), item.get("invoiceAmount"), item.get("invoiceSourceFile"), item.get("invoiceSourceRowNo"), item.get("settlementFile"), item.get("settlementPeriod"), item.get("settlementAmount"), item.get("settlementConfidence"), item.get("cashflowDate"), item.get("receivedAmount"), item.get("cashflowSourceFile"), item.get("cashflowSourceRowNo"), item.get("status"), item.get("abnormalReason"), item.get("abnormalStage"), item.get("settlementFileId"), "是" if item.get("manualCheckRequired") else "否"])
        _style_status_row(ws, ws.max_row, item.get("status"))

    ws = wb.create_sheet("异常项清单")
    ws.append(headers)
    _style_header(ws, 1)
    for item in reconciliation["items"]:
        if item.get("manualCheckRequired") or item.get("status") == "发票已红冲":
            ws.append([item.get("customerName"), item.get("invoiceNo"), item.get("invoiceDate"), item.get("revenueMonth"), item.get("invoiceAmount"), item.get("invoiceSourceFile"), item.get("invoiceSourceRowNo"), item.get("settlementFile"), item.get("settlementPeriod"), item.get("settlementAmount"), item.get("settlementConfidence"), item.get("cashflowDate"), item.get("receivedAmount"), item.get("cashflowSourceFile"), item.get("cashflowSourceRowNo"), item.get("status"), item.get("abnormalReason"), item.get("abnormalStage"), item.get("settlementFileId"), "是" if item.get("manualCheckRequired") else "否"])
            _style_status_row(ws, ws.max_row, item.get("status"))

    ws = wb.create_sheet("解析文件列表")
    ws.append(["文件ID", "文件名", "文件类型", "解析状态", "解析行数", "有效行数", "置信度", "失败或警告原因"])
    _style_header(ws, 1)
    for file in files:
        ws.append([file.get("fileId"), file.get("fileName"), file.get("fileType"), file.get("parseStatus"), file.get("parsedRows"), file.get("validRows"), file.get("confidence"), file.get("errorReason")])

    for sheet in wb.worksheets:
        for col in sheet.columns:
            letter = col[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) + 3 for cell in col), 36)
    wb.save(path)


def _sheet(wb, preferred: str | None):
    if preferred and preferred in wb.sheetnames:
        return wb[preferred]
    return wb[wb.sheetnames[0]]


def _find_sheet_header(wb, required: list[str], preferred: str | None):
    sheets = []
    if preferred and preferred in wb.sheetnames:
        sheets.append(wb[preferred])
    sheets.extend(wb[name] for name in wb.sheetnames if not preferred or name != preferred)
    last_error: ValueError | None = None
    for ws in sheets:
        try:
            header_row, headers = _find_header(ws, required)
            return ws, header_row, headers
        except ValueError as exc:
            last_error = exc
    raise last_error or ValueError(f"未找到表头：{', '.join(required)}")


def _find_header(ws, required: list[str]) -> tuple[int, dict[str, int]]:
    for row_no in range(1, min(ws.max_row, 20) + 1):
        values = [_text(ws.cell(row_no, col).value) for col in range(1, ws.max_column + 1)]
        if all(item in values for item in required):
            return row_no, {value: index + 1 for index, value in enumerate(values) if value}
    raise ValueError(f"未找到表头：{', '.join(required)}")


def _row_dict(ws, row_no: int, headers: dict[str, int]) -> dict[str, Any]:
    return {key: _cell_value(ws.cell(row_no, col).value) for key, col in headers.items()}


def _extract_settlement_text(path: Path) -> tuple[str, str, str | None]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(path, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"# {ws.title}")
            for row in ws.iter_rows(values_only=True):
                values = [_text(value) for value in row]
                if any(values):
                    lines.append(" | ".join(values))
        return "\n".join(lines), "success", None
    if suffix == ".docx":
        text = _docx_text(path)
        return (text, "success", None) if text.strip() else ("", "failed", "DOCX 未提取到有效文本")
    if suffix == ".pdf":
        text = _pdf_text(path)
        if _pdf_text_has_settlement_detail(text):
            return text, "success", None
        ocr_text, ocr_status, ocr_reason = _ocr_pdf(path)
        if ocr_status == "success":
            return _merge_text(text, ocr_text), "success", None
        if text.strip():
            return text, "success", None
        return ocr_text, ocr_status, ocr_reason
    if suffix in {".txt", ".csv"}:
        return path.read_text(encoding="utf-8", errors="ignore"), "success", None
    if suffix in {".png", ".jpg", ".jpeg", ".webp"}:
        return _ocr_image(path)
    return "", "failed", "不支持的结算单文件类型"


def _pdf_text(path: Path) -> str:
    try:
        import fitz

        with fitz.open(path) as doc:
            return "\n".join(page.get_text() for page in doc)
    except Exception:
        return ""


def _pdf_text_has_settlement_detail(text: str) -> bool:
    compact = re.sub(r"\s+", "", text or "")
    if len(compact) < 80:
        return False
    has_amount = bool(re.search(r"\d+(?:,\d{3})*(?:\.\d{1,2})?", compact))
    has_settlement_terms = any(term in compact for term in ["合计", "金额", "服务内容", "结算周期", "技术服务费"])
    return has_amount and has_settlement_terms


def _merge_text(*parts: str) -> str:
    seen: set[str] = set()
    lines: list[str] = []
    for part in parts:
        for line in (part or "").splitlines():
            normalized = line.strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            lines.append(normalized)
    return "\n".join(lines)


# ── OCR 引擎（基于 rapidocr-onnxruntime） ──────────────────────────────────────

_ocr_engine = None


def _get_ocr():
    global _ocr_engine
    if _ocr_engine is None:
        from rapidocr_onnxruntime import RapidOCR

        _ocr_engine = RapidOCR()
    return _ocr_engine


def _ocr_image(path: Path) -> tuple[str, str, str | None]:
    """对单张图片执行 OCR，返回 (text, status, error_message)。"""
    try:
        ocr = _get_ocr()
        result, elapse = ocr(str(path))
        if result is None:
            return "", "needs_ocr", "OCR 未识别到有效文本"
        lines = [item[1] for item in result]
        return "\n".join(lines), "success", None
    except ImportError:
        return "", "needs_ocr", "OCR 引擎未安装，请安装 rapidocr-onnxruntime"
    except Exception as exc:
        return "", "failed", f"OCR 识别失败：{exc}"


def _ocr_pdf(path: Path) -> tuple[str, str, str | None]:
    """将 PDF 每页渲染为图片后执行 OCR，返回 (text, status, error_message)。"""
    try:
        import fitz
    except ImportError:
        return "", "needs_ocr", "缺少 pymupdf，无法渲染扫描 PDF"

    try:
        ocr = _get_ocr()
    except ImportError:
        return "", "needs_ocr", "OCR 引擎未安装，请安装 rapidocr-onnxruntime"

    try:
        all_lines: list[str] = []
        with fitz.open(path) as doc:
            pages = list(doc)
            if not pages:
                return "", "failed", "PDF 文件没有可识别页面"

            for page_idx, page in enumerate(pages, start=1):
                pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                tmp = path.with_name(f".ocr_tmp_{path.stem}_p{page_idx}.png")
                pixmap.save(tmp)
                try:
                    result, _elapse = ocr(str(tmp))
                    if result:
                        all_lines.extend(item[1] for item in result)
                    all_lines.append("")
                finally:
                    tmp.unlink(missing_ok=True)

        text = "\n".join(all_lines).strip()
        if text:
            return text, "success", None
        return "", "needs_ocr", "PDF 经 OCR 后未提取到有效文本"
    except Exception as exc:
        return "", "failed", f"PDF OCR 失败：{exc}"


def _docx_text(path: Path) -> str:
    try:
        with zipfile.ZipFile(path) as zf:
            xml = zf.read("word/document.xml")
        root = ElementTree.fromstring(xml)
        ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
        return "\n".join(node.text for node in root.findall(".//w:t", ns) if node.text)
    except Exception:
        return ""


def _find_match(
    source: dict[str, Any],
    candidates: list[dict[str, Any]],
    amount_key: str,
    used: set[int],
    alternate_names: list[Any] | None = None,
    candidate_label: str = "",
) -> tuple[int | None, dict[str, Any] | None, str | None, str | None]:
    source_amount = Decimal(str(source.get("invoiceAmount") or 0))
    amount_candidates: list[tuple[int, dict[str, Any]]] = []
    for index, item in enumerate(candidates):
        if index in used:
            continue
        amount = Decimal(str(item.get(amount_key) or 0))
        if abs(amount - source_amount) <= MONEY_TOLERANCE:
            amount_candidates.append((index, item))
    if not amount_candidates:
        return None, None, None, f"未找到同金额{candidate_label}记录"

    source_names = [source.get("customerName"), *(alternate_names or [])]
    ranked: dict[CompanyMatchRelation, list[tuple[int, dict[str, Any]]]] = {
        CompanyMatchRelation.EXACT: [],
        CompanyMatchRelation.SAME_MAIN_ENTITY: [],
    }
    for index, item in amount_candidates:
        relations = [compare_company_names(name, item.get("customerName")) for name in source_names if name]
        relation = CompanyMatchRelation.EXACT if CompanyMatchRelation.EXACT in relations else (
            CompanyMatchRelation.SAME_MAIN_ENTITY if CompanyMatchRelation.SAME_MAIN_ENTITY in relations else None
        )
        if relation:
            ranked[relation].append((index, item))

    for relation in (CompanyMatchRelation.EXACT, CompanyMatchRelation.SAME_MAIN_ENTITY):
        matches = ranked[relation]
        if len(matches) == 1:
            index, item = matches[0]
            return index, item, relation.value, None
        if len(matches) > 1:
            return None, None, None, f"同金额、同主体存在多个{candidate_label}候选，需人工确认"
    return None, None, None, f"存在同金额{candidate_label}记录，但公司主体不一致"


def _standard_json_for_file(job_dir: Path, file_id: str, file_type: str) -> Any:
    if file_type == "invoice_excel":
        return _read_json(job_dir / "parsed" / "invoices.json", [])
    if file_type == "cashflow_excel":
        return _read_json(job_dir / "parsed" / "cashflows.json", [])
    settlements = _read_json(job_dir / "parsed" / "settlements.json", [])
    return [item for item in settlements if item.get("fileId") == file_id]


def _job_dir(job_id: str) -> Path:
    configured = Path(get_settings().income_reconciliation_storage_dir) / "income-reconciliation" / "jobs"
    root = JOB_ROOT if JOB_ROOT != Path("storage/income-reconciliation/jobs") else configured
    return root / job_id


def _require_job(job_id: str) -> Path:
    path = _job_dir(job_id)
    if not path.exists():
        raise AppError("job_not_found", "收入核对任务不存在", 404)
    return path


def _save_upload(file: UploadFile, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    file.file.seek(0)
    with target.open("wb") as out:
        shutil.copyfileobj(file.file, out)


def _safe_name(name: str) -> str:
    return Path(name).name.replace("/", "_").replace("\\", "_")


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(f".{path.name}.{uuid.uuid4().hex}.tmp")
    tmp.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(path)


def _read_json(path: Path, default: Any = None) -> Any:
    if not path or not path.exists() or not path.is_file():
        return default
    try:
        content = path.read_text(encoding="utf-8")
        if not content.strip():
            return default
        return json.loads(content)
    except json.JSONDecodeError:
        logger.warning("income_reconciliation_json_read_failed path=%s", path)
        return default


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def _read_text(path: Path) -> str:
    if not path or not path.exists() or not path.is_file():
        return ""
    return path.read_text(encoding="utf-8", errors="ignore")


def _reset_progress_events(job_dir: Path) -> None:
    path = job_dir / "progress.jsonl"
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("", encoding="utf-8")


def _append_progress_event(job_dir: Path, event_type: str, message: str, **payload: Any) -> dict[str, Any]:
    path = job_dir / "progress.jsonl"
    event = {
        "seq": len(get_progress_events_for_dir(job_dir)) + 1,
        "type": event_type,
        "message": message,
        "createdAt": datetime.now().isoformat(),
        **payload,
    }
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as output:
        output.write(json.dumps(event, ensure_ascii=False) + "\n")
    return event


def get_progress_events_for_dir(job_dir: Path) -> list[dict[str, Any]]:
    path = job_dir / "progress.jsonl"
    if not path.exists():
        return []
    events = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        try:
            events.append(json.loads(line))
        except json.JSONDecodeError:
            logger.warning("income_reconciliation_progress_event_invalid path=%s line=%s", path, line[:200])
    return events


def _initial_file_statuses(metadata: dict[str, Any]) -> list[dict[str, Any]]:
    files = [
        _file_status("file_invoice", metadata["invoiceFile"], "invoice_excel", "pending", 0, 0, 0, None),
        _file_status("file_cashflow", metadata["cashflowFile"], "cashflow_excel", "pending", 0, 0, 0, None),
    ]
    files.extend(
        _file_status(f"file_settlement_{index:03d}", name, _settlement_type(Path(name)), "pending", 0, 0, 0, None)
        for index, name in enumerate(metadata.get("settlementFiles", []), start=1)
    )
    return files


def _replace_file_status(files: list[dict[str, Any]], status: dict[str, Any]) -> None:
    for index, item in enumerate(files):
        if item.get("fileId") == status.get("fileId"):
            files[index] = status
            return
    files.append(status)


def _file_status(file_id: str, file_name: str, file_type: str, status: str, parsed: int, valid: int, confidence: float, reason: str | None) -> dict[str, Any]:
    return {"fileId": file_id, "fileName": file_name, "fileType": file_type, "parseStatus": status, "parsedRows": parsed, "validRows": valid, "confidence": confidence, "errorReason": reason}


def _settlement_type(path: Path) -> str:
    suffix = path.suffix.lower()
    return {".xlsx": "settlement_excel", ".xlsm": "settlement_excel", ".pdf": "settlement_pdf", ".docx": "settlement_docx"}.get(suffix, "settlement_file")


def _cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()[:10]
    if isinstance(value, Decimal):
        return _float(value)
    return value


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _money(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value).replace(",", "").replace("￥", "").replace("¥", "").replace("CNY", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _float(value: Decimal | None) -> float | None:
    if value is None:
        return None
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    return float(value.quantize(Decimal("0.01")))


def _date_text(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _text(value)
    if not text:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}", text):
        return text[:10]
    match = re.search(r"(20\d{2})[年/-](\d{1,2})[月/-](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return text


def _month_text(value: Any) -> str | None:
    text = _text(value)
    if not text:
        return None
    if re.match(r"^\d{4}-\d{2}$", text):
        return text
    match = re.search(r"(20\d{2})[年/-](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}"
    return text[:7] if re.match(r"^\d{4}-\d{2}-", text) else text


def _norm_name(value: Any) -> str:
    return normalize_company_name(value)


def _json_from_text(text: str) -> Any:
    if not isinstance(text, str):
        return None
    match = re.search(r"\{.*\}", text, re.S)
    if not match:
        return None
    return json.loads(match.group(0))


def _abnormal_stage(status: str) -> str | None:
    return {"金额异常": "金额匹配", "已确认未到账": "资料缺失", "未确认已到账": "资料缺失", "资料缺失待确认": "资料缺失", "发票已红冲": "字段校验"}.get(status)


def _style_header(ws, row: int) -> None:
    fill = PatternFill("solid", fgColor="EAF3F6")
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = fill


def _style_status_row(ws, row: int, status: str | None) -> None:
    colors = {"金额异常": "FDE2E2", "已确认未到账": "FFF4CC", "未确认已到账": "DDEEFF", "资料缺失待确认": "FFF4CC", "发票已红冲": "EEEEEE"}
    color = colors.get(status)
    if not color:
        return
    fill = PatternFill("solid", fgColor=color)
    for cell in ws[row]:
        cell.fill = fill
