from datetime import date
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database.base import Base
from app.database.session import get_db
from app.main import app
from app.modules.cashflow import models as cashflow_models  # noqa
from app.modules.data_import import models as import_models  # noqa
from app.modules.income_reconciliation.service import _extract_settlement_text
from tests.unit.test_parser import workbook_bytes


PROJECT_ROOT = Path(__file__).resolve().parents[3]


@pytest.fixture
def client():
    engine = create_engine("sqlite+pysqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(engine)
    Session = sessionmaker(engine, expire_on_commit=False)
    def override():
        with Session() as db: yield db
    app.dependency_overrides[get_db] = override
    with TestClient(app) as test_client: yield test_client
    app.dependency_overrides.clear()


def test_import_publish_and_dashboard(client):
    response = client.post("/api/v1/imports", data={"templateVersion": "V0.1", "dataPeriodStart": "2026-05-01", "dataPeriodEnd": "2026-07-31"}, files={"file": ("sample.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert response.status_code == 201, response.text
    batch = response.json()["data"]
    assert batch["errorRows"] == 0
    published = client.post(f"/api/v1/imports/{batch['id']}:publish", json={"version": batch["version"], "reviewNote": "测试确认"})
    assert published.status_code == 200, published.text
    overview = client.get("/api/v1/cockpit/overview", params={"asOfDate": "2026-07-02"})
    assert overview.status_code == 200, overview.text
    assert overview.json()["data"]["metrics"]["availableCash"] == "3000000.00"
    assert overview.json()["data"]["metrics"]["overdueAmount"] == "1000000.00"
    risks = client.get("/api/v1/receivable-risks", params={"asOfDate": "2026-07-02"})
    assert risks.json()["data"][0]["riskLevel"] == "red"


def test_uniform_error_shape(client):
    response = client.get("/api/v1/cockpit/overview")
    assert response.status_code == 404
    assert response.json()["error"]["code"] == "published_batch_not_found"


def test_deterministic_cockpit_modules(client):
    response = client.post("/api/v1/imports", data={"templateVersion": "V0.1", "dataPeriodStart": "2026-05-01", "dataPeriodEnd": "2026-09-30"}, files={"file": ("modules.xlsx", workbook_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    batch = response.json()["data"]
    published = client.post(f"/api/v1/imports/{batch['id']}:publish", json={"version": batch["version"]})
    assert published.status_code == 200, published.text

    forecast = client.get("/api/v1/cashflow-forecasts", params={"asOfDate": "2026-07-02"})
    assert forecast.status_code == 200, forecast.text
    assert len(forecast.json()["data"]["daily"]) == 90
    assert {item["days"] for item in forecast.json()["data"]["windows"]} == {30, 60, 90}

    risk_top3 = client.get("/api/v1/receivable-risks/top3", params={"asOfDate": "2026-07-02"})
    assert risk_top3.status_code == 200
    assert risk_top3.json()["data"]["redCount"] == 1
    assert risk_top3.json()["data"]["items"][0]["reasonCodes"] == ["overdue_30_days", "overdue_large_amount"]

    payments = client.get("/api/v1/payment-recommendations/top3", params={"asOfDate": "2026-07-02"})
    assert payments.status_code == 200, payments.text
    payment = payments.json()["data"]["items"][0]
    assert payment["decision"] == "pay"
    assert payment["aiExplanation"] is None
    payment_explanation = client.post(f"/api/v1/payment-recommendations/{payment['id']}/ai-explanation")
    assert payment_explanation.status_code == 200, payment_explanation.text
    assert payment_explanation.json()["data"]["cached"] is False
    assert "一句话结论" in payment_explanation.json()["data"]["aiExplanation"]
    cached_explanation = client.post(f"/api/v1/payment-recommendations/{payment['id']}/ai-explanation")
    assert cached_explanation.status_code == 200, cached_explanation.text
    assert cached_explanation.json()["data"]["cached"] is True

    events = client.get("/api/v1/decision-events", params={"asOfDate": "2026-07-02"})
    assert events.status_code == 200, events.text
    event = events.json()["data"][0]
    assert event["eventType"] == "receivable_risk"
    explanation = client.get(f"/api/v1/decision-events/{event['id']}/ai-explanation:stream")
    assert explanation.status_code == 200
    assert explanation.headers["content-type"].startswith("text/event-stream")
    assert "一句话结论" in explanation.text
    assert '"degraded": true' in explanation.text
    decided = client.post(f"/api/v1/decision-events/{event['id']}:decide", json={"option": "escalate_collection", "note": "升级处理", "version": event["version"]})
    assert decided.status_code == 200, decided.text
    assert decided.json()["data"]["status"] == "decided"
    refreshed = client.get("/api/v1/decision-events", params={"asOfDate": "2026-07-02"})
    assert refreshed.status_code == 200
    assert refreshed.json()["pagination"]["total"] == 0


def test_income_reconciliation_demo_flow(client, tmp_path, monkeypatch):
    monkeypatch.setattr("app.modules.income_reconciliation.service.JOB_ROOT", tmp_path)
    data_dir = PROJECT_ROOT / "test-data"
    with (
        (data_dir / "测试数据26.4-5月开票明细.xlsx").open("rb") as invoice,
        (data_dir / "测试数据2026年4月-5月服务收支明细.xlsx").open("rb") as cashflow,
        (data_dir / "5.15-正邦测试数据-10614.54.xlsx").open("rb") as settlement,
    ):
        response = client.post(
            "/api/v1/income-reconciliation/jobs",
            data={"period_start": "2026-04", "period_end": "2026-05"},
            files=[
                ("invoice_file", ("invoice.xlsx", invoice, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                ("cashflow_file", ("cashflow.xlsx", cashflow, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
                ("settlement_files", ("settlement.xlsx", settlement, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ],
        )
    assert response.status_code == 201, response.text
    job = response.json()["data"]
    assert job["status"] == "uploaded"

    parsed = client.post(f"/api/v1/income-reconciliation/jobs/{job['jobId']}/parse")
    assert parsed.status_code == 200, parsed.text
    parsed_data = parsed.json()["data"]
    assert parsed_data["status"] == "parsed"
    assert len(parsed_data["files"]) == 3
    events = client.get(f"/api/v1/income-reconciliation/jobs/{job['jobId']}/events.json")
    assert events.status_code == 200, events.text
    event_types = [event["type"] for event in events.json()["data"]["events"]]
    assert "job_started" in event_types
    assert "job_done" in event_types

    generated = client.post(f"/api/v1/income-reconciliation/jobs/{job['jobId']}/generate")
    assert generated.status_code == 200, generated.text
    generated_data = generated.json()["data"]
    assert generated_data["status"] == "generated"
    assert generated_data["summary"]["confirmedRevenue"] > 0
    assert generated_data["downloadUrl"].endswith("/download")

    download = client.get(f"/api/v1/income-reconciliation/jobs/{job['jobId']}/download")
    assert download.status_code == 200, download.text
    output = tmp_path / "download.xlsx"
    output.write_bytes(download.content)
    workbook = load_workbook(output, read_only=True)
    assert {"老板卡片", "收入链路核对表", "异常项清单", "解析文件列表"}.issubset(workbook.sheetnames)


def test_settlement_pdf_uses_ocr_when_text_layer_is_incomplete():
    pdf_path = PROJECT_ROOT / "test-data" / "紫金陈小说公司5月结算单（5.13）.pdf"

    text, status, reason = _extract_settlement_text(pdf_path)
    assert status == "success", reason
    assert "# 第 1 页" in text
    assert "服务内容" in text
    assert "32038.95" in text
